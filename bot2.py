#!/usr/bin/env python3
"""
BOT 2 - DUAL MODE TRADING
Mode 1: SPAM VOLUME - Market in/out immediately on Hyperliquid
Mode 2: DELTA HEDGE - Hold BTC 10x leverage across HL + Lighter for OI farming
"""

import os
import sys
import time
import json
import asyncio
import logging
import random
from dataclasses import dataclass
from typing import Optional, Dict, Literal, Tuple
from decimal import Decimal
from dotenv import load_dotenv

# Load env vars
load_dotenv()
if os.path.exists('.env.lighter'):
    load_dotenv('.env.lighter')

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(name)-12s | %(levelname)-5s | %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger("Bot2")

@dataclass
class TradeConfig:
    symbol: str = "BTC"
    size_usd: float = 500.0  # XXX$ - điều chỉnh theo ý bạn
    
    # Mode 1: Spam Volume (có thể để leverage cao vì vào ra ngay)
    spam_leverage: int = 10  # Bẩy cho mode spam
    spam_interval_sec: float = 10.0  # ⏱️ Delay giữa các lệnh (tăng lên 10s để tránh rate limit)
    spam_rounds: int = 10  # 🔁 Số lần spam (chỉnh ở đây)
    spam_size_range: Tuple[float, float] = (100.0, 400.0)  # 💰 Random size $100-$400 mỗi lệnh
    
    # Mode 2: Delta Hedge (bẩy thấp an toàn vì hold lâu)
    hedge_leverage: int = 4
    
    # Mode 2: Delta Hedge - OI Farming Settings
    hl_paradex_ratio: float = 1.0  # 1:1 hedge
    funding_threshold: float = 0.0001  # Đóng khi funding rate < 0.01%
    profit_target_usd: float = 5.0  # Chốt lờ khi đủ $
    
    # ⚙️ OI HOLD SETTINGS (Tuỳ chỉnh ở đây)
    hedge_hold_hours: float = 8.0       # ⏱️ Thởi gian hold mỗi cycle (4h-12h)
    hedge_auto_reenter: bool = True      # 🔄 Auto vào lệnh mới sau khi đóng
    hedge_cycles: int = -1               # 🔁 Số cycle chạy (-1 = vô hạn)
    hedge_rest_hours: float = 0.5        # 💤 Nghỉ bao lâu giữa các cycle (30 phút)
    hedge_exit_on_fee: bool = True       # ✅ Đóng khi đủ gỡ phí HL (ưu tiên an toàn)


class HyperliquidTrader:
    """Wrapper cho Hyperliquid - dùng maker limit order (ALO) để nhận rebate"""
    
    MAKER_TIMEOUT_SEC = 5  # Chờ fill tối đa 5 giây
    
    def __init__(self):
        from valiant_exchange import ValiantHyperliquidExchange
        self.exchange = ValiantHyperliquidExchange()
        self.symbol = "BTC"
        
    def get_position(self) -> Optional[Dict]:
        positions = self.exchange.get_positions()
        for p in positions:
            if p["symbol"] == self.symbol:
                return p
        return None
    
    def get_bbo(self) -> Optional[Dict]:
        """Lấy Best Bid/Offer (BBO) từ L2 orderbook"""
        try:
            book = self.exchange.info.l2_snapshot(self.symbol)
            if book and "levels" in book:
                levels = book["levels"]
                if len(levels) >= 2 and len(levels[0]) > 0 and len(levels[1]) > 0:
                    best_bid = float(levels[0][0]["px"])
                    best_ask = float(levels[1][0]["px"])
                    return {"bid": best_bid, "ask": best_ask, "spread": best_ask - best_bid}
            return None
        except Exception as e:
            logger.debug(f"Failed to get BBO: {e}")
            return None
    
    def _maker_order(self, side: str, size_btc: float, price: float = None) -> Dict:
        """
        Đặt maker limit order (ALO) @ best bid/ask.
        Dùng ValiantExchange.limit_order (đã test OK) với giá từ BBO.
        """
        is_buy = side.lower() in ("long", "buy")
        
        # Lấy BBO tươi nhất
        bbo = self.get_bbo()
        if bbo:
            limit_price = bbo["bid"] if is_buy else bbo["ask"]
        else:
            # Fallback: mid - 2 tick
            TICK = 2.0
            limit_price = (price or 70000) - TICK if is_buy else (price or 70000) + TICK
        
        return self._maker_order_with_price(side, size_btc, limit_price, bbo)
    
    def _maker_order_with_price(self, side: str, size_btc: float, limit_price: float, bbo: dict = None) -> Dict:
        """Đặt ALO với giá cụ thể (để retry)"""
        is_buy = side.lower() in ("long", "buy")
        hl_side = "buy" if is_buy else "sell"
        
        if bbo:
            logger.info(f"  📋 ALO {hl_side.upper()} {size_btc:.6f} BTC @ ${limit_price:.1f} (bid=${bbo['bid']:.1f} ask=${bbo['ask']:.1f})")
        else:
            logger.info(f"  📋 ALO {hl_side.upper()} {size_btc:.6f} BTC @ ${limit_price:.1f}")
        
        # Dùng ValiantExchange.limit_order (đã handle round + margin check)
        result = self.exchange.limit_order(self.symbol, hl_side, size_btc, limit_price, tif="Alo")
        
        return result
    
    def maker_open(self, side: str, size_usd: float, wait_fill: bool = True, max_wait_sec: int = 60) -> Dict:
        """
        Mở position bằng maker order (ALO)
        
        Flow:
        1. Đặt ALO limit order @ best bid/ask
        2. Nếu wait_fill=True: Chờ cho đến khi fill hoàn toàn (max 60s)
        3. Nếu wait_fill=False: Trả về ngay (resting)
        4. Nếu timeout → cancel và báo lỗi (không fallback market)
        
        Returns: {"status": "filled", ...} hoặc {"status": "error", "error": ...}
        """
        try:
            price = self.exchange.get_mid_price(self.symbol)
            if price is None:
                logger.error("Cannot get price")
                return {"error": "Cannot get price", "status": "error"}
            
            size_btc = size_usd / price
            is_buy = side.lower() in ("long", "buy")
            
            # 1. Đặt maker order
            result = self._maker_order(side, size_btc, price)
            limit_price = None  # Will be set from BBO in _maker_order
            
            if result.get("error"):
                logger.error(f"  Maker order failed: {result.get('error')}")
                return {"error": result.get("error"), "status": "error"}
            
            # 2. Parse response
            if result.get("status") == "ok":
                statuses = result.get("response", {}).get("data", {}).get("statuses", [])
                if statuses:
                    status = statuses[0]
                    
                    # Filled ngay → great!
                    if "filled" in status:
                        fill_info = status["filled"]
                        filled_size = float(fill_info.get("totalSz", size_btc))
                        logger.info(f"  ✅ Maker FILLED immediately! Size={filled_size:.6f} BTC, Rebate earned 💰")
                        return {"status": "filled", "filled_size": filled_size, "response": result}
                    
                    # Resting trên book → chờ fill nếu cần
                    if "resting" in status:
                        oid = status["resting"]["oid"]
                        
                        if not wait_fill:
                            logger.info(f"  ⏳ Maker order resting (oid={oid}), not waiting")
                            return {"status": "resting", "oid": oid, "response": result}
                        
                        logger.info(f"  ⏳ Maker order resting (oid={oid}), waiting for fill (max {max_wait_sec}s)...")
                        
                        # Poll chờ fill cho đến khi có position
                        for i in range(max_wait_sec):
                            time.sleep(1)
                            pos = self.get_position()
                            pos_size = abs(float(pos.get("size", 0))) if pos else 0
                            
                            if pos_size > 0.00001:
                                logger.info(f"  ✅ Maker FILLED after {i+1}s! Size={pos_size:.6f} BTC, Rebate earned 💰")
                                return {"status": "filled", "filled_size": pos_size, "waited_sec": i+1}
                            
                            # Log mỗi 5s
                            if (i + 1) % 5 == 0:
                                logger.info(f"  ⏳ Still waiting... {i+1}s elapsed")
                        
                        # Timeout → cancel và báo lỗi (KHÔNG fallback market)
                        logger.warning(f"  ⏰ Maker timeout after {max_wait_sec}s, cancelling order...")
                        try:
                            self.exchange.cancel_order(self.symbol, oid)
                            time.sleep(0.5)
                        except Exception as e:
                            logger.debug(f"  Cancel error (may already filled): {e}")
                        
                        # Check lại position sau khi cancel
                        pos = self.get_position()
                        pos_size = abs(float(pos.get("size", 0))) if pos else 0
                        
                        if pos_size > 0.00001:
                            logger.info(f"  ✅ Position detected after cancel! Size={pos_size:.6f} BTC")
                            return {"status": "filled", "filled_size": pos_size, "note": "filled_during_cancel"}
                        
                        logger.error(f"  ⏰ Maker timeout after {max_wait_sec}s, not filled")
                        return {"error": f"Maker order timeout after {max_wait_sec}s, not filled", "status": "error"}
                    
                    # Error trong status - retry with adjusted price
                    if "error" in status:
                        error_msg = status['error']
                        logger.error(f"  ALO rejected: {error_msg}")
                        
                        # If "Post only order would have immediately matched", adjust price and retry
                        if "immediately matched" in error_msg.lower() or "post only" in error_msg.lower():
                            logger.info(f"  Retrying ALO with adjusted price...")
                            # Wait a moment and get fresh BBO
                            time.sleep(0.5)
                            fresh_bbo = self.get_bbo()
                            if fresh_bbo:
                                # Adjust price by 1 tick away from mid
                                TICK = 1.0
                                old_price = fresh_bbo["bid"] if is_buy else fresh_bbo["ask"]
                                if is_buy:
                                    new_price = fresh_bbo["bid"] - TICK
                                else:
                                    new_price = fresh_bbo["ask"] + TICK
                                logger.info(f"  Retrying ALO @ ${new_price:.1f} (was ${old_price:.1f})")
                                # Recalculate size for new price
                                new_size_btc = size_usd / new_price
                                return self._maker_order_with_price(side, new_size_btc, new_price)
                        
                        return {"error": status["error"], "status": "error"}
            
            # Không parse được → error
            logger.error(f"  Unexpected response: {result}")
            return {"error": "Unexpected response", "status": "error", "response": result}
            
        except Exception as e:
            logger.error(f"Maker open failed: {e}")
            return {"error": str(e), "status": "error"}
    
    def maker_close(self, size_usd: float = None) -> Dict:
        """Đóng position bằng maker order + fallback market"""
        try:
            pos = self.get_position()
            if not pos or abs(float(pos.get("size", 0))) < 0.00001:
                return {"status": "no_position"}
            
            size = abs(float(pos.get("size", 0)))
            side = "short" if float(pos.get("size", 0)) > 0 else "long"
            price = self.exchange.get_mid_price(self.symbol)
            if not price:
                price = float(pos.get("entry_price", 70000))
            
            # Đặt maker close - BBO tươi + gửi trực tiếp (no delay)
            hl_side = "sell" if side == "short" else "buy"
            
            bbo = self.get_bbo()
            if bbo:
                limit_price = bbo["ask"] if hl_side == "sell" else bbo["bid"]
                logger.info(f"  📋 ALO close {hl_side.upper()} {size:.6f} BTC @ ${limit_price:.1f} (bid=${bbo['bid']:.1f} ask=${bbo['ask']:.1f})")
            else:
                TICK = 2.0
                limit_price = price + TICK if hl_side == "sell" else price - TICK
                logger.info(f"  📋 ALO close {hl_side.upper()} {size:.6f} BTC @ ${limit_price:.1f} (no BBO)")
            
            result = self.exchange.limit_order(self.symbol, hl_side, size, limit_price, tif="Alo")
            
            if result.get("status") == "ok":
                statuses = result.get("response", {}).get("data", {}).get("statuses", [])
                if statuses:
                    status = statuses[0]
                    
                    if "filled" in status:
                        logger.info(f"  ✅ Maker close FILLED! Rebate earned 💰")
                        return result
                    
                    if "resting" in status:
                        oid = status["resting"]["oid"]
                        logger.info(f"  ⏳ Maker close resting, waiting {self.MAKER_TIMEOUT_SEC}s...")
                        
                        for i in range(self.MAKER_TIMEOUT_SEC):
                            time.sleep(1)
                            pos = self.get_position()
                            if not pos or abs(float(pos.get("size", 0))) < 0.00001:
                                logger.info(f"  ✅ Maker close FILLED after {i+1}s! Rebate earned 💰")
                                return result
                        
                        # Timeout → cancel + market close
                        logger.warning(f"  ⏰ Timeout, cancelling + market close")
                        self.exchange.cancel_order(self.symbol, oid)
                        time.sleep(0.5)
            
            # Fallback: market close
            return self.close_position()
            
        except Exception as e:
            logger.error(f"Maker close failed: {e}, falling back to market")
            return self.close_position()
    
    def market_order(self, side: Literal["long", "short"], size_usd: float) -> Dict:
        """Market order trên HL (fallback, taker fee)"""
        try:
            price = self.exchange.get_mid_price(self.symbol)
            if price is None:
                logger.error("Cannot get price, skipping this round")
                return {"error": "Cannot get price"}
            size_btc = size_usd / price
            
            result = self.exchange.market_order(self.symbol, side, size_btc)
            logger.info(f"HL {side.upper()} {size_usd}$ @ {price:.2f} = {size_btc:.6f} BTC")
            return result
        except Exception as e:
            logger.error(f"HL order failed: {e}")
            return {"error": str(e)}
    
    def close_position(self) -> Dict:
        """Đóng toàn bộ position (market, fallback)"""
        try:
            result = self.exchange.market_close(self.symbol)
            logger.info(f"HL CLOSE position: {result}")
            return result
        except Exception as e:
            logger.error(f"HL close failed: {e}")
            return {"error": str(e)}
    
    def get_funding_rate(self) -> float:
        """Lấy funding rate hiện tại"""
        try:
            funding = self.exchange.get_funding_rate(self.symbol)
            return float(funding.get("funding_rate", 0)) if isinstance(funding, dict) else float(funding)
        except:
            return 0.0


class LighterTraderWrapper:
    """Wrapper cho Lighter - SDK mode (thay Paradex) - Đã fix Windows"""
    
    def __init__(self):
        self.trader = None
        self.connected = False
        self.account = None
        self.last_position = None  # Cache position để close không cần API call thêm
        
        try:
            from lighter_trader_sdk import LighterSDKTrader
            self.trader = LighterSDKTrader()
            self.account = f"Account {self.trader.account_index}"
            logger.info(f"Lighter SDK initialized: {self.account}")
        except Exception as e:
            logger.error(f"❌ Lighter SDK init: {e}")
            logger.warning("Chạy mode HL-only (không có hedge)")
    
    async def connect(self):
        """Khởi tạo kết nối async"""
        if self.trader:
            await self.trader.connect()
            self.connected = True
            logger.info("Lighter connected!")
    
    async def get_funding_rate(self) -> float:
        if not self.connected:
            return 0.0
        try:
            return await self.trader.get_funding_rate()
        except:
            return 0.0
    
    async def market_order(self, side: Literal["long", "short"], size_usd: float, price: float = None) -> Dict:
        if not self.connected:
            return {"status": "error", "msg": "Not connected"}
        try:
            return await self.trader.market_order(side, size_usd, price=price)
        except Exception as e:
            return {"status": "error", "msg": str(e)}
    
    async def close_position(self) -> Dict:
        if not self.connected:
            return {"status": "error", "msg": "Not connected"}
        try:
            return await self.trader.close_position()
        except Exception as e:
            return {"status": "error", "msg": str(e)}
    
    async def close_position_direct(self, price: float = None) -> Dict:
        """Close dùng cached position data - chỉ 1 API call (không cần get_position)"""
        if not self.connected:
            return {"status": "error", "msg": "Not connected"}
        try:
            if self.last_position:
                return await self.trader.close_position_direct(self.last_position, price=price)
            else:
                return await self.trader.close_position()
        except Exception as e:
            return {"status": "error", "msg": str(e)}
    
    async def get_position(self) -> Optional[Dict]:
        if not self.connected:
            return None
        try:
            pos = await self.trader.get_position()
            if pos:
                self.last_position = pos  # Cache
            return pos
        except:
            return None


class Bot2:
    """
    Bot 2 chế độ:
    - MODE 1: Spam volume HL (vào-ra ngay)
    - MODE 2: Delta hedge HL+Paradex (hold lâu, farm OI)
    """
    
    def __init__(self, config: TradeConfig = None):
        self.config = config or TradeConfig()
        self.hl = HyperliquidTrader()
        self.lighter = LighterTraderWrapper()
        self.mode: Literal["spam", "hedge", "off"] = "off"
        
        # Stats
        self.stats = {
            "spam_trades": 0,
            "spam_volume": 0.0,
            "hedge_opened": None,
            "hedge_pnl": 0.0,
        }
        
        # Cycle reports cho Excel export
        self.cycle_reports = []
    
    async def init_lighter(self):
        """Khởi tạo kết nối Lighter (async)"""
        if self.lighter.trader:
            try:
                await self.lighter.connect()
                
                # Verify connection works by making a test API call
                logger.info("Verifying Lighter API access...")
                test_price = await self.lighter.trader.get_mid_price()
                if test_price is None:
                    logger.error("Lighter API test failed: Cannot get price (403 error)")
                    logger.error("Switching to HL-only mode")
                    self.lighter.connected = False
                else:
                    logger.info(f"Lighter connected! BTC price: ${test_price:,.2f}")
            except Exception as e:
                logger.warning(f"Lighter connect failed: {e}")
                logger.warning("Mode 2 will run HL-only")
                self.lighter.connected = False
    
    async def safety_check_on_startup(self):
        """Check nếu có position cũ còn sót (bot crash trước đó) → cảnh báo"""
        logger.info("🔍 Safety check: checking for stale positions...")
        
        hl_pos = self.hl.get_position()
        hl_size = abs(float(hl_pos.get("size", 0))) if hl_pos else 0
        
        lighter_size = 0
        if self.lighter.connected:
            lighter_pos = await self.lighter.get_position()
            lighter_size = lighter_pos.get("size", 0) if lighter_pos else 0
        
        if hl_size > 0.00001 or lighter_size > 0.00001:
            logger.warning(f"⚠️  STALE POSITIONS FOUND!")
            logger.warning(f"  HL: {hl_size:.8f} BTC")
            logger.warning(f"  Lighter: {lighter_size:.8f} BTC")
            logger.warning(f"  Có thể do bot crash trước đó.")
            logger.warning(f"  Hãy close thủ công trước khi chạy hedge mode.")
            logger.warning(f"  Hoặc gõ 'close_all' để đóng tất cả.")
            return True
        else:
            logger.info("  ✅ No stale positions. Safe to trade.")
            return False
    
    async def get_position_size_btc(self, exchange: str) -> float:
        """Lấy position size BTC từ 1 exchange (async)"""
        try:
            if exchange == "hl":
                pos = self.hl.get_position()
                if pos:
                    return abs(float(pos.get("size", 0)))
            elif exchange == "lighter":
                pos = await self.lighter.get_position()
                if pos:
                    # Lighter trả về size đã là absolute value
                    return float(pos.get("size", 0))
        except Exception as e:
            logger.debug(f"get_position_size_btc error: {e}")
        return 0.0
    
    async def perfect_delta_entry(self, side_hl: str, side_lighter: str, 
                                   usd_size: float, max_slippage: float = 0.001) -> bool:
        """
        Vào lệnh PERFECT DELTA - Đảm bảo size BTC 2 bên bằng nhau tuyệt đối
        
        Strategy:
        1. Lấy giá HL làm reference
        2. Tính size BTC chính xác = usd_size / price_hl
        3. Vào HL với size USD
        4. Vào Lighter với cùng size BTC (tính lại USD nếu giá khác)
        
        Args:
            side_hl: "long" or "short" cho HL
            side_lighter: "long" or "short" cho Lighter
            usd_size: Size USD (VD: 100, 200, 400)
            max_slippage: Max slippage cho phép (0.1%)
        
        Returns: True nếu thành công
        """
        if not self.lighter.connected:
            logger.error("Lighter not connected, cannot do perfect delta")
            return False
        
        try:
            # 1. Lấy giá cả 2 sàn và kiểm tra chênh lệch (quan trọng cho delta chuẩn)
            price_hl = self.hl.exchange.get_mid_price(self.config.symbol)
            price_lighter = await self.lighter.trader.get_mid_price()
            
            if not price_hl and not price_lighter:
                logger.error("Cannot get prices from both exchanges")
                return False
            if not price_hl:
                logger.error("Cannot get Hyperliquid price")
                return False
            if not price_lighter:
                logger.error("Cannot get Lighter price - API key may be invalid (403 error)")
                logger.error("Check: 1) LIGHTER_API_PRIVATE_KEYS format, 2) Account index matches, 3) API key not expired")
                return False
            
            # Kiểm tra price divergence
            price_diff = abs(price_hl - price_lighter) / price_hl
            max_price_diff = 0.001  # 0.1% max cho delta chuẩn
            
            logger.info(f"PRICE CHECK (quan trọng cho delta):")
            logger.info(f"  HL: ${price_hl:,.2f}")
            logger.info(f"  Lighter: ${price_lighter:,.2f}")
            logger.info(f"  Diff: {price_diff*100:.3f}% (max cho phép: 0.1%)")
            
            if price_diff > max_price_diff:
                logger.warning(f"CAUTION: Price diff > 0.1% - Delta sẽ không chuẩn!")
                logger.warning(f"Nếu BTC move mạnh, PnL 2 bên sẽ khác nhau")
            else:
                logger.info(f"Price OK - Good for delta neutral")
            
            # 2. Tính size BTC chính xác - dùng giá trung bình cho chuẩn
            avg_price = (price_hl + price_lighter) / 2
            size_btc_exact = usd_size / avg_price
            size_satoshi = int(size_btc_exact * 10**8)
            size_btc_final = size_satoshi / 10**8
            
            logger.info(f"PERFECT DELTA ENTRY")
            logger.info(f"  Avg price: ${avg_price:,.2f}")
            logger.info(f"  Target size: {size_btc_final:.8f} BTC ({size_satoshi} satoshi)")
            
            # Kiểm tra balance Lighter TRƯỚC (dùng hết API read ở đây)
            lighter_balance = await self.lighter.trader.get_balance()
            logger.info(f"  Lighter balance: ${lighter_balance:.2f}")
            
            # ⚠️ CHECK LIGHTER STALE POSITION (từ lần trước close fail)
            lighter_existing = await self.lighter.get_position()
            if lighter_existing and lighter_existing.get("size", 0) > 0.00001:
                logger.warning(f"  ⚠️ Lighter has existing position: {lighter_existing.get('size')} BTC!")
                logger.warning(f"  Closing stale position first...")
                await asyncio.sleep(15)  # Rate limit
                await self.lighter.close_position()
                await asyncio.sleep(15)  # Rate limit
                # Verify closed
                lighter_check = await self.lighter.get_position()
                if lighter_check and lighter_check.get("size", 0) > 0.00001:
                    logger.error(f"  Cannot close stale Lighter position! Aborting.")
                    return False
                logger.info(f"  ✅ Stale position closed.")
            
            usd_size_lighter = size_btc_final * price_lighter
            margin_needed = usd_size_lighter / 10  # leverage x10
            if lighter_balance < margin_needed * 1.1:
                logger.error(f"Insufficient balance! Need ~${margin_needed:.2f} margin, have ${lighter_balance:.2f}")
                return False
            
            # 3. VÀO LỆNH - HL (maker) CHỜ FILL rồi gửi Lighter NGAY
            logger.info(f"[1] Entering HL {side_hl.upper()} (maker order → rebate 💰)...")
            logger.info(f"  Waiting for HL ALO to fill completely before sending Lighter...")
            
            hl_order = self.hl.maker_open(side_hl, usd_size, wait_fill=True, max_wait_sec=60)
            
            if hl_order.get("status") != "filled":
                error_msg = hl_order.get("error", "unknown error")
                logger.error(f"HL entry failed or timeout: {error_msg}")
                logger.error("Not sending Lighter order - no HL position!")
                return False
            
            # ✅ HL đã filled - lấy actual filled size
            hl_filled_size = hl_order.get("filled_size", size_btc_final)
            logger.info(f"  ✅ HL filled: {hl_filled_size:.8f} BTC (rebate earned 💰)")
            
            # Tính USD cần cho Lighter (same BTC size) - GỬI NGAY, không chờ rate limit
            usd_size_lighter = hl_filled_size * price_lighter
            
            logger.info(f"[2] Entering Lighter {side_lighter.upper()} (HL already filled)...")
            logger.info(f"  Target: {hl_filled_size:.8f} BTC (same as HL)")
            logger.info(f"  USD needed: ${usd_size_lighter:,.2f} (@ ${price_lighter:,.2f})")
            
            lighter_order = await self.lighter.market_order(
                side_lighter, 
                usd_size_lighter,
                price=price_lighter
            )
            
            # Check lỗi
            is_error = (lighter_order.get("error") or 
                       lighter_order.get("msg") or 
                       lighter_order.get("status") == "error")
            if is_error:
                err_msg = lighter_order.get("error") or lighter_order.get("msg") or "unknown"
                logger.error(f"Lighter order failed: {err_msg}")
                logger.error("HL position is filled but Lighter failed!")
                logger.error("Closing HL to avoid naked exposure...")
                self.hl.close_position()
                return False
            
            logger.info("  ✅ Lighter order sent!")
            logger.info("  ✅ Both sides entered!")
            
            # Chờ 15s rồi verify Lighter position (rate limit)
            await asyncio.sleep(15)
            lighter_pos = await self.lighter.get_position()
            hl_pos = self.hl.get_position()  # Lấy HL position cho entry price
            
            # 5. Verify perfect delta
            if lighter_pos:
                lighter_size = lighter_pos.get("size", 0)
                mismatch = abs(hl_filled_size - lighter_size)
                mismatch_pct = mismatch / hl_filled_size if hl_filled_size > 0 else 0
                
                # Lấy entry prices
                hl_entry = float(hl_pos.get("entry_price") or price_hl) if hl_pos else price_hl
                lighter_entry = lighter_pos.get("entry_price", price_lighter)
                entry_diff = abs(hl_entry - lighter_entry)
                entry_diff_pct = entry_diff / hl_entry if hl_entry > 0 else 0
                
                logger.info(f"PERFECT DELTA RESULT:")
                logger.info(f"  SIZE:")
                logger.info(f"    HL:      {hl_filled_size:.8f} BTC ({side_hl.upper()})")
                logger.info(f"    Lighter: {lighter_size:.8f} BTC ({side_lighter.upper()})")
                logger.info(f"    Mismatch: {mismatch:.8f} BTC ({mismatch_pct*100:.4f}%)")
                logger.info(f"  ENTRY PRICE (quan trọng cho delta PnL):")
                logger.info(f"    HL: ${hl_entry:,.2f}")
                logger.info(f"    Lighter: ${lighter_entry:,.2f}")
                logger.info(f"    Diff: ${entry_diff:.2f} ({entry_diff_pct*100:.3f}%)")
                
                # Tính PnL impact nếu BTC move 1%
                btc_move_pct = 0.01
                hl_pnl_if_move = hl_filled_size * hl_entry * btc_move_pct * (-1 if side_hl == "short" else 1)
                lighter_pnl_if_move = lighter_size * lighter_entry * btc_move_pct * (-1 if side_lighter == "short" else 1)
                net_pnl_if_move = hl_pnl_if_move + lighter_pnl_if_move
                
                logger.info(f"  PnL SIMULATION (BTC +1%):")
                logger.info(f"    HL: ${hl_pnl_if_move:+.4f}")
                logger.info(f"    Lighter: ${lighter_pnl_if_move:+.4f}")
                logger.info(f"    Net: ${net_pnl_if_move:+.4f} (càng gần 0 càng chuẩn)")
                
                # Status
                size_ok = mismatch_pct < 0.0005
                price_ok = entry_diff_pct < 0.001
                
                if size_ok and price_ok:
                    logger.info(f"  STATUS: PERFECT DELTA ✅ (Size + Price)")
                elif size_ok:
                    logger.warning(f"  STATUS: OK Size, BAD Price ({entry_diff_pct*100:.3f}%)")
                elif price_ok:
                    logger.warning(f"  STATUS: BAD Size, OK Price")
                else:
                    logger.error(f"  STATUS: BAD - Cần xem xét lại!")
                
                return True
            
            return True
            
        except Exception as e:
            logger.error(f"Perfect delta entry failed: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    async def verify_and_fix_delta(self, tolerance: float = 0.001) -> bool:
        """
        Kiểm tra delta và fix nếu không hoàn hảo
        Returns: True nếu đã fix
        """
        try:
            hl_size = await self.get_position_size_btc("hl")
            lighter_size = await self.get_position_size_btc("lighter")
            
            if hl_size == 0 or lighter_size == 0:
                return False
            
            mismatch = abs(hl_size - lighter_size)
            mismatch_pct = mismatch / max(hl_size, lighter_size)
            
            if mismatch_pct <= tolerance:
                logger.info(f"Delta OK: {mismatch_pct*100:.4f}% mismatch")
                return False
            
            logger.warning(f"Delta imperfect: {mismatch_pct*100:.2f}% mismatch")
            logger.warning(f"HL: {hl_size:.8f} BTC, Lighter: {lighter_size:.8f} BTC")
            
            # Chỉ cảnh báo, KHÔNG tự close (close 1 bên khi bên kia vẫn mở = naked exposure)
            # Delta luôn 0.0000% từ entry nên trường hợp này hiếm khi xảy ra
            logger.warning(f"Delta mismatch detected but NOT auto-fixing (too risky)")
            logger.warning(f"Consider manual intervention if mismatch persists")
            return False
            
        except Exception as e:
            logger.error(f"Verify delta failed: {e}")
            return False
        
    def set_leverage(self, leverage: int = None):
        """Set leverage trên HL"""
        lev = leverage or self.config.hedge_leverage
        try:
            self.hl.exchange.set_leverage(self.config.symbol, lev)
            logger.info(f"Set {self.config.symbol} leverage = {lev}x")
        except Exception as e:
            logger.warning(f"Leverage set failed: {e}")
    
    def _calc_liquidation(self, entry_price: float, side: str, leverage: int) -> float:
        """Tính giá liquidation"""
        # HL: Maint margin ~0.5%, liq khi margin còn ~0.5%
        # Approx: liq_price = entry * (1 ± 1/leverage) cho cross margin
        liq_distance = 0.95 / leverage  # 95% / leverage (conservative)
        
        if side == "long":
            return entry_price * (1 - liq_distance)
        else:
            return entry_price * (1 + liq_distance)
    
    def _funding_pnl_calc(self, size_usd: float, funding_rate: float, hours: float) -> float:
        """Tính lãi/lỗ từ funding
        
        Args:
            size_usd: Position size
            funding_rate: % mỗi 8h (vd: 0.0001 = 0.01%)
            hours: Thởi gian hold
        
        Returns: USD PnL từ funding
        """
        periods = hours / 8  # Funding mỗi 8h
        return size_usd * funding_rate * periods
    
    # ═══════════════════════════════════════════════════════
    # MODE 1: SPAM VOLUME
    # ═══════════════════════════════════════════════════════
    
    async def run_spam_mode(self):
        """
        Mode 1: Spam market orders để tạo volume
        Vào long/short rồi đóng ngay lập tức
        Size random $100-$400 mỗi lệnh
        """
        min_size, max_size = self.config.spam_size_range
        
        logger.info(f"\n{'='*50}")
        logger.info("MODE 1: SPAM VOLUME ACTIVATED")
        logger.info(f"Size: RANDOM ${min_size:.0f}-${max_size:.0f}, Leverage: {self.config.spam_leverage}x")
        logger.info(f"Interval: {self.config.spam_interval_sec}s, Rounds: {self.config.spam_rounds}")
        logger.info(f"{'='*50}\n")
        
        self.mode = "spam"
        self.set_leverage(self.config.spam_leverage)
        
        for i in range(self.config.spam_rounds):
            if self.mode != "spam":
                break
            
            # Random size cho mỗi lệnh
            size_usd = random.uniform(min_size, max_size)
            size_usd = round(size_usd, 2)  # Làm tròn 2 số thập phân
                
            logger.info(f"\n--- Spam Round {i+1}/{self.config.spam_rounds} | Size: ${size_usd:.2f} ---")
            
            # Random long/short để tạo volume 2 chiều
            side = "long" if i % 2 == 0 else "short"
            
            # 1. Vào lệnh
            price = self.hl.exchange.get_mid_price(self.config.symbol)
            if price is None:
                logger.error("Cannot get price, skipping this round")
                await asyncio.sleep(5)
                continue
            
            liq_price = self._calc_liquidation(price, side, self.config.spam_leverage)
            liq_dist = abs(price - liq_price) / price * 100
            
            logger.info(f"⚠️  {side.upper()} ${size_usd:.2f} @ {price:.0f} | Liq: {liq_price:.0f} ({liq_dist:.1f}%)")
            
            entry = self.hl.market_order(side, size_usd)
            if entry.get("error"):
                logger.error(f"Entry failed: {entry['error']}")
                await asyncio.sleep(2)
                continue
            
            self.stats["spam_trades"] += 1
            self.stats["spam_volume"] += size_usd
            
            # 2. Chờ 2 giây rồi đóng (đảm bảo order filled)
            await asyncio.sleep(2)
            
            # 3. Đóng position (thử 3 lần nếu fail)
            for attempt in range(3):
                close = self.hl.close_position()
                if not close.get("error"):
                    break
                logger.warning(f"Close attempt {attempt+1} failed: {close['error']}")
                await asyncio.sleep(2)
            else:
                logger.error(f"Close failed after 3 attempts")
            
            # 4. Chờ interval
            logger.info(f"Spam complete. Total volume: {self.stats['spam_volume']:.0f}$")
            await asyncio.sleep(self.config.spam_interval_sec)
        
        logger.info(f"\nSpam mode complete. Total: {self.stats['spam_trades']} trades, {self.stats['spam_volume']:.0f}$ volume")
    
    # ═══════════════════════════════════════════════════════
    # MODE 2: DELTA HEDGE
    # ═══════════════════════════════════════════════════════
    
    async def run_hedge_mode(self):
        """
        Mode 2: Delta neutral hedge giữa HL và Paradex
        - OI Farming: Hold 4-12h rồi re-enter
        - Tự động chạy nhiều cycle
        """
        # Check Lighter API ready
        if not self.lighter.connected:
            logger.warning("⚠️ Lighter API chưa config! Chạy HL-only mode.")
            logger.warning("Set env: LIGHTER_API_KEY để dùng full delta hedge")
        
        self.mode = "hedge"
        cycle_count = 0
        
        logger.info(f"\n{'='*60}")
        logger.info("MODE 2: OI DELTA HEDGE - MULTI CYCLE")
        logger.info(f"Exchanges: Hyperliquid + Lighter (Arbitrum)")
        logger.info(f"Size: {self.config.size_usd}$ per leg | Leverage: {self.config.hedge_leverage}x")
        logger.info(f"⏱️  Hold time: {self.config.hedge_hold_hours}h/cycle")
        logger.info(f"🔄 Auto re-enter: {self.config.hedge_auto_reenter} | Max cycles: {self.config.hedge_cycles if self.config.hedge_cycles > 0 else '∞'}")
        logger.info(f"💤 Rest between: {self.config.hedge_rest_hours}h")
        logger.info(f"🔗 Lighter: {'✓ CONNECTED' if self.lighter.connected else '✗ DISCONNECTED'}")
        logger.info(f"{'='*60}\n")
        
        while self.mode == "hedge":
            # Check max cycles
            if self.config.hedge_cycles > 0 and cycle_count >= self.config.hedge_cycles:
                logger.info(f"✅ Completed {cycle_count} cycles. Stopping hedge mode.")
                break
            
            cycle_count += 1
            logger.info(f"\n{'='*60}")
            logger.info(f"🚀 CYCLE #{cycle_count} STARTING...")
            logger.info(f"{'='*60}")
            
            # Run single hedge cycle
            should_continue = await self._run_single_hedge_cycle()
            
            if not should_continue or self.mode != "hedge":
                break
            
            # Rest before next cycle
            if self.config.hedge_auto_reenter:
                logger.info(f"\n💤 Resting {self.config.hedge_rest_hours}h before next cycle...")
                rest_sec = self.config.hedge_rest_hours * 3600
                await asyncio.sleep(rest_sec)
        
        self.mode = "off"
        logger.info("\n🏁 Hedge mode completed.")
        
        # Export báo cáo Excel
        if self.cycle_reports:
            self._export_report()
    
    async def _run_single_hedge_cycle(self) -> bool:
        """
        Chạy 1 cycle hedge: vào lệnh → hold X giờ → đóng
        
        Returns: True nếu nên tiếp tục cycle tiếp theo
        """
        self.set_leverage(self.config.hedge_leverage)
        
        # Check funding rates để chọn hướng
        hl_funding = self.hl.get_funding_rate()
        lighter_funding = await self.lighter.get_funding_rate()
        
        logger.info(f"HL Funding: {hl_funding*100:.4f}%")
        logger.info(f"Lighter Funding: {lighter_funding*100:.4f}%")
        
        # Tính funding PnL dự kiến cho hold time
        hl_funding_pnl = self._funding_pnl_calc(self.config.size_usd, hl_funding, self.config.hedge_hold_hours)
        lighter_funding_pnl = self._funding_pnl_calc(self.config.size_usd, lighter_funding, self.config.hedge_hold_hours)
        total_pnl = hl_funding_pnl + lighter_funding_pnl
        logger.info(f"💰 Est. funding PnL this cycle ({self.config.hedge_hold_hours}h): {total_pnl:.3f}$")
        
        # ⚠️ LIQUIDATION WARNING
        price = self.hl.exchange.get_mid_price(self.config.symbol)
        liq_long = self._calc_liquidation(price, "long", self.config.hedge_leverage)
        liq_short = self._calc_liquidation(price, "short", self.config.hedge_leverage)
        logger.info(f"⚠️  Liq prices: Long {liq_long:.0f} | Short {liq_short:.0f} | Current: {price:.0f}")
        
        # Chiến lược: Long sàn có funding âm (nhận tiền), Short sàn có funding dương (trả ít)
        if hl_funding < lighter_funding:
            hl_side = "long"
            lighter_side = "short"
            logger.info("Strategy: LONG HL + SHORT Lighter")
        else:
            hl_side = "short"
            lighter_side = "long"
            logger.info("Strategy: SHORT HL + LONG Lighter")
        
        # PERFECT DELTA ENTRY
        if self.lighter.connected:
            entry_success = await self.perfect_delta_entry(
                side_hl=hl_side,
                side_lighter=lighter_side,
                usd_size=self.config.size_usd
            )
            if not entry_success:
                logger.error("Perfect delta entry failed!")
                return False
        else:
            # HL only mode
            logger.info(f"[1] Opening HL {hl_side} (HL-only mode)...")
            hl_order = self.hl.maker_open(hl_side, self.config.size_usd)
            if hl_order.get("error"):
                logger.error(f"HL entry failed: {hl_order['error']}")
                return False
        
        self.stats["hedge_opened"] = time.time()
        entry_time = time.strftime("%Y-%m-%d %H:%M:%S")
        entry_price = price
        
        # Lấy entry prices từ positions
        hl_entry_price = price
        lighter_entry_price = price
        hl_pos = self.hl.get_position()
        if hl_pos:
            hl_entry_price = float(hl_pos.get("entry_price") or price)
        
        # 3. Monitor và hold
        logger.info(f"\n[3] Monitoring for {self.config.hedge_hold_hours}h...")
        close_reason = await self._monitor_hedge_cycle()
        
        # Lấy PnL trước khi đóng
        hl_pnl = 0.0
        hl_pos = self.hl.get_position()
        if hl_pos:
            hl_pnl = float(hl_pos.get("unrealized_pnl") or 0)
        
        # Lighter PnL ≈ ngược lại HL PnL (delta neutral)
        lighter_pnl = -hl_pnl
        
        # Lấy Lighter entry price từ cached position
        lighter_pos = self.lighter.last_position
        if lighter_pos:
            lighter_entry_price = float(lighter_pos.get("entry_price", 0) or price)
        else:
            lighter_entry_price = price  # fallback
        
        exit_price = self.hl.exchange.get_mid_price(self.config.symbol) or price
        
        # 4. Đóng vị thế
        fee = self._calc_hl_fee(self.config.size_usd)
        net_after_fee = (hl_pnl + lighter_pnl) - fee
        logger.info(f"\n[4] Closing positions (reason: {close_reason})...")
        logger.info(f"  HL PnL: ${hl_pnl:.4f} | Lighter PnL (est): ${lighter_pnl:.4f} | Fee: ${fee:.4f} | Net: ${net_after_fee:.4f}")
        
        # Lưu thời gian trước khi close (close sẽ reset hedge_opened = None)
        opened_at = self.stats["hedge_opened"] or time.time()
        
        await self._close_hedge()
        
        exit_time = time.strftime("%Y-%m-%d %H:%M:%S")
        hold_minutes = (time.time() - opened_at) / 60
        
        # Lưu cycle report (wrapped để không crash nếu thiếu data)
        try:
            self.cycle_reports.append({
                "cycle": len(self.cycle_reports) + 1,
                "entry_time": entry_time,
                "exit_time": exit_time,
                "hold_minutes": round(hold_minutes, 1),
                "hl_side": hl_side,
                "lighter_side": lighter_side,
                "size_usd": self.config.size_usd,
                "leverage": self.config.hedge_leverage,
                "entry_price": round(entry_price or 0, 2),
                "exit_price": round(exit_price or 0, 2),
                "hl_entry": round(hl_entry_price or 0, 2),
                "lighter_entry": round(lighter_entry_price or 0, 2),
                "hl_funding": round(hl_funding * 100, 4),
                "lighter_funding": round(lighter_funding * 100, 4),
                "hl_pnl": round(hl_pnl or 0, 4),
                "lighter_pnl": round(lighter_pnl or 0, 4),
                "net_pnl": round((hl_pnl or 0) + (lighter_pnl or 0), 4),
                "fee_hl": round(self._calc_hl_fee(self.config.size_usd), 4),
                "close_reason": close_reason,
            })
        except Exception as e:
            logger.error(f"Failed to save cycle report: {e}")
        
        return self.config.hedge_auto_reenter
    
    def _calc_hl_fee(self, size_usd: float) -> float:
        """Tính phí HL cho 1 round trip (open + close)
        
        Hyperliquid fees:
        - Maker (ALO success): REBATE -0.015% → -$0.045 per $150 round trip
        - Taker (market/ALO fail): 0.045% → +$0.135 per $150 round trip
        
        Tính taker (worst case) để report chính xác.
        Nếu maker fill thì thực tế sẽ tốt hơn report.
        """
        hl_taker_fee = 0.00045  # 0.045%
        round_trip_fee = hl_taker_fee * 2
        return size_usd * round_trip_fee
    
    def _export_report(self):
        """Xuất báo cáo Excel sau khi chạy xong tất cả cycles"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            
            # ═══ Sheet 1: SUMMARY ═══
            ws = wb.active
            ws.title = "Summary"
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
            header_fill = PatternFill("solid", fgColor="2F5496")
            data_font = Font(size=10, name="Arial")
            money_green = Font(size=10, name="Arial", color="008000")
            money_red = Font(size=10, name="Arial", color="FF0000")
            border = Border(
                bottom=Side(style="thin", color="D9D9D9"),
            )
            center = Alignment(horizontal="center", vertical="center")
            
            # Title
            ws.merge_cells("A1:H1")
            ws["A1"] = "BOT 2 - DELTA HEDGE REPORT"
            ws["A1"].font = Font(bold=True, size=14, name="Arial", color="2F5496")
            ws["A1"].alignment = Alignment(horizontal="center")
            
            # Summary info
            ws["A3"] = "Generated"
            ws["B3"] = time.strftime("%Y-%m-%d %H:%M:%S")
            ws["A4"] = "Total Cycles"
            ws["B4"] = len(self.cycle_reports)
            ws["A5"] = "Size/Leg"
            ws["B5"] = f"${self.config.size_usd:.0f}"
            ws["A6"] = "Leverage"
            ws["B6"] = f"{self.config.hedge_leverage}x"
            ws["A7"] = "Hold Time"
            ws["B7"] = f"{self.config.hedge_hold_hours}h"
            
            for row in range(3, 8):
                ws[f"A{row}"].font = Font(bold=True, size=10, name="Arial")
                ws[f"B{row}"].font = data_font
            
            # Totals
            total_hl_pnl = sum(c["hl_pnl"] for c in self.cycle_reports)
            total_lighter_pnl = sum(c["lighter_pnl"] for c in self.cycle_reports)
            total_net_pnl = sum(c["net_pnl"] for c in self.cycle_reports)
            total_fee = sum(c["fee_hl"] for c in self.cycle_reports)
            total_hold = sum(c["hold_minutes"] for c in self.cycle_reports)
            
            ws["D3"] = "HL PnL"
            ws["E3"] = total_hl_pnl
            ws["E3"].number_format = '$#,##0.0000'
            ws["E3"].font = money_green if total_hl_pnl >= 0 else money_red
            ws["D4"] = "Lighter PnL (est)"
            ws["E4"] = total_lighter_pnl
            ws["E4"].number_format = '$#,##0.0000'
            ws["E4"].font = money_green if total_lighter_pnl >= 0 else money_red
            ws["D5"] = "Net PnL"
            ws["E5"] = total_net_pnl
            ws["E5"].number_format = '$#,##0.0000'
            ws["E5"].font = money_green if total_net_pnl >= 0 else money_red
            ws["D6"] = "Total HL Fee"
            ws["E6"] = total_fee
            ws["E6"].number_format = '$#,##0.0000'
            ws["D7"] = "Net After Fee"
            ws["E7"] = total_net_pnl - total_fee
            ws["E7"].number_format = '$#,##0.0000'
            ws["E7"].font = Font(bold=True, size=11, name="Arial", color="008000" if (total_net_pnl - total_fee) >= 0 else "FF0000")
            ws["D8"] = "Total Hold"
            ws["E8"] = f"{total_hold:.0f} min"
            
            for row in range(3, 9):
                ws[f"D{row}"].font = Font(bold=True, size=10, name="Arial")
            
            # ═══ Sheet 2: CYCLES DETAIL ═══
            ws2 = wb.create_sheet("Cycles")
            
            headers = [
                "Cycle", "Entry Time", "Exit Time", "Hold (min)",
                "HL Side", "Lighter Side", "Size ($)", "Leverage",
                "Entry Price", "Exit Price", "Price Change",
                "HL Entry", "Lighter Entry", "Entry Diff",
                "HL Funding%", "Lighter Funding%",
                "HL PnL ($)", "Lighter PnL ($)", "Net PnL ($)", 
                "HL Fee ($)", "Net After Fee ($)", "Close Reason"
            ]
            
            # Write headers
            for col, h in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
            
            # Write data
            for i, c in enumerate(self.cycle_reports, 2):
                price_change = c["exit_price"] - c["entry_price"]
                entry_diff = abs(c["hl_entry"] - c["lighter_entry"])
                net_after_fee = c["net_pnl"] - c["fee_hl"]
                
                row_data = [
                    c["cycle"], c["entry_time"], c["exit_time"], c["hold_minutes"],
                    c["hl_side"].upper(), c["lighter_side"].upper(), c["size_usd"], c["leverage"],
                    c["entry_price"], c["exit_price"], round(price_change, 2),
                    c["hl_entry"], c["lighter_entry"], round(entry_diff, 2),
                    c["hl_funding"], c["lighter_funding"],
                    c["hl_pnl"], c["lighter_pnl"], c["net_pnl"],
                    c["fee_hl"], round(net_after_fee, 4), c["close_reason"]
                ]
                
                for col, val in enumerate(row_data, 1):
                    cell = ws2.cell(row=i, column=col, value=val)
                    cell.font = data_font
                    cell.border = border
                    if col in (17, 18, 19):  # PnL columns
                        cell.number_format = '$#,##0.0000'
                        cell.font = money_green if val >= 0 else money_red
                    elif col == 21:  # Net After Fee
                        cell.number_format = '$#,##0.0000'
                        cell.font = money_green if val >= 0 else money_red
                    elif col in (9, 10, 12, 13):  # Price columns
                        cell.number_format = '#,##0.00'
            
            # Auto column widths
            for ws_sheet in [ws, ws2]:
                for col in range(1, ws_sheet.max_column + 1):
                    max_len = 0
                    for row in range(1, ws_sheet.max_row + 1):
                        cell = ws_sheet.cell(row=row, column=col)
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws_sheet.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 25)
            
            # Save
            filename = f"hedge_report_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            logger.info(f"\n📊 Report saved: {filename}")
            logger.info(f"   Cycles: {len(self.cycle_reports)} | HL PnL: ${total_hl_pnl:.4f} | Lighter PnL: ${total_lighter_pnl:.4f} | Fee: ${total_fee:.4f} | Net: ${total_net_pnl - total_fee:.4f}")
            
        except Exception as e:
            logger.error(f"Failed to export report: {e}")
            import traceback
            traceback.print_exc()
    
    async def _monitor_hedge_cycle(self) -> str:
        """
        Giám sát 1 hedge cycle cho đến khi:
        - Đủ hold time (target)
        - Đủ lợi nhuận
        - Đủ fee HL (gỡ lại phí)
        - Funding đảo chiều
        - Liq quá gần
        
        Returns: lý do đóng lệnh
        """
        start_time = time.time()
        check_interval = 60  # Check mỗi phút
        
        # Tính fee cần gỡ
        fee_target = self._calc_hl_fee(self.config.size_usd)
        logger.info(f"💸 Target: Gỡ lại phí HL = ${fee_target:.3f} (round trip)")
        
        while self.mode == "hedge":
            elapsed = time.time() - start_time
            elapsed_hours = elapsed / 3600
            remaining = self.config.hedge_hold_hours - elapsed_hours
            
            # Check position HL
            hl_pos = self.hl.get_position()
            hl_funding = self.hl.get_funding_rate()
            
            if hl_pos:
                # Ensure numeric values (convert from string if needed, handle None)
                pnl = float(hl_pos.get("unrealized_pnl") or 0)
                entry = float(hl_pos.get("entry_price") or 0)
                side = str(hl_pos.get("side", "LONG")).lower()
                liq = float(hl_pos.get("liquidation_price") or 0)
                
                # Calc liq distance
                if liq and entry and liq > 0:
                    dist = abs(entry - liq) / entry * 100
                else:
                    # Không có liq price (chưa lấy được) → dùng leverage ước tính
                    # Tránh dist=0 gây panic close sai
                    dist = (0.95 / self.config.hedge_leverage) * 100  # fallback an toàn
                    logger.debug(f"No liq price from exchange, using estimated dist: {dist:.1f}%")
                
                progress = (elapsed_hours / self.config.hedge_hold_hours) * 100
                fee_status = "✓" if abs(pnl) >= fee_target else "..."
                net = pnl - fee_target  # Net after fee (negative fee = profit)
                logger.info(f"[{elapsed_hours:.1f}h/{self.config.hedge_hold_hours}h {progress:.0f}%] HL PnL: {pnl:.3f}$ | Fee: ${fee_target:.3f} {fee_status} | Net: ${net:.3f} | LiqDist: {dist:.1f}%")
                
                # CẢNH BÁO LIQUIDATION
                if dist < 5:
                    logger.error(f"DANGER! Liq only {dist:.1f}% away!")
                elif dist < 8:
                    logger.warning(f"WARNING! Liq {dist:.1f}% away")
                
                # Check delta mismatch mỗi 5 phút
                if int(elapsed) % 300 < 60 and self.lighter.connected:
                    await self.verify_and_fix_delta()
                
                # Điều kiện đóng SỚM (theo thứ tự ưu tiên):
                
                # 1. Liq quá gần - đóng ngay để an toàn
                if dist < 4:
                    return f"liquidation_risk ({dist:.1f}%)"
                
                # 2. Đủ lợi nhuận target - đóng sớm
                if abs(pnl) >= self.config.profit_target_usd:
                    return f"profit_target ({pnl:.2f}$)"
            
            # 3. Đủ thời gian hold → check fee rồi đóng
            if elapsed_hours >= self.config.hedge_hold_hours:
                # Nếu bật exit_on_fee: đã gỡ được phí → exit clean
                if self.config.hedge_exit_on_fee and abs(pnl) >= fee_target:
                    return f"hold_complete + fee_recovered (PnL: ${pnl:.3f} >= Fee: ${fee_target:.3f})"
                # Chưa gỡ phí nhưng hết time → vẫn exit
                return f"hold_time_complete ({self.config.hedge_hold_hours}h, PnL: ${pnl:.3f})"
            
            await asyncio.sleep(check_interval)
        
        return "mode_stopped"
    
    async def _close_hedge(self) -> Dict:
        """Đóng cả 2 vị thế hedge - tối ưu để close gần nhau nhất"""
        logger.info("\n[5] Closing hedge positions...")
        
        # Lấy giá hiện tại cho Lighter close (cache position đã có từ verify_delta)
        exit_price = self.hl.exchange.get_mid_price(self.config.symbol)
        
        # Nếu chưa có cached position, get trước rồi chờ 15s
        if not self.lighter.last_position and self.lighter.connected:
            await self.lighter.get_position()  # Cache position
            logger.info("  ⏳ Waiting 15s for Lighter rate limit...")
            await asyncio.sleep(15)
        else:
            # Đã có cached position, chỉ cần chờ 15s kể từ API call cuối
            logger.info("  ⏳ Waiting 15s for Lighter rate limit...")
            await asyncio.sleep(15)
        
        # CLOSE CẢ 2 BACK-TO-BACK (không delay giữa)
        # 1. HL close (maker → rebate)
        hl_close = self.hl.maker_close()
        logger.info(f"  HL closed: {'OK' if not hl_close.get('error') else hl_close.get('error')}")
        
        # 2. Lighter close ngay sau HL (dùng cached position, 1 API call only)
        lighter_close = {"status": "not_connected"}
        if self.lighter.connected:
            lighter_close = await self.lighter.close_position_direct(price=exit_price)
            
            # Nếu nonce error → chờ 15s rồi retry bằng close_position thường
            if lighter_close.get("status") == "error":
                err = lighter_close.get("msg") or lighter_close.get("error") or ""
                if "nonce" in str(err).lower():
                    logger.warning(f"  Lighter nonce error, waiting 15s and retrying...")
                    await asyncio.sleep(15)
                    lighter_close = await self.lighter.close_position()
                else:
                    logger.warning(f"  Lighter close error: {err}, waiting 15s and retrying...")
                    await asyncio.sleep(15)
                    lighter_close = await self.lighter.close_position()
            
            status = lighter_close.get("status", "unknown")
            logger.info(f"  Lighter closed: {status}")
        
        # Verify HL
        await asyncio.sleep(1)
        hl_pos = self.hl.get_position()
        if hl_pos and abs(float(hl_pos.get("size") or 0)) > 0.00001:
            logger.warning(f"  HL still open: {hl_pos.get('size')} BTC, retrying...")
            self.hl.close_position()
        
        # Clear cached position
        self.lighter.last_position = None
        self.stats["hedge_opened"] = None
        
        return {
            "hl": hl_close,
            "lighter": lighter_close,
            "success": not hl_close.get("error")
        }
    
    # ═══════════════════════════════════════════════════════
    # CLI INTERFACE
    # ═══════════════════════════════════════════════════════
    
    async def run(self):
        """Main loop với interactive commands"""
        # Init Lighter connection
        await self.init_lighter()
        
        # Safety check for stale positions
        has_stale = await self.safety_check_on_startup()
        
        # Show config summary
        cycles_str = "inf" if self.config.hedge_cycles < 0 else str(self.config.hedge_cycles)
        print(f"""
============================================================
                    BOT 2 - DUAL MODE
============================================================
  Capital: ~$110 USDC | Symbol: BTC-USD
============================================================
  Mode 1: SPAM VOLUME
    - Size: ${self.config.size_usd:.0f} | Leverage: {self.config.spam_leverage}x | Margin: ~${self.config.size_usd/self.config.spam_leverage:.0f}
    - Rounds: {self.config.spam_rounds} | Interval: {self.config.spam_interval_sec}s
============================================================
  Mode 2: OI DELTA HEDGE (HL + Lighter)
    - Size: ${self.config.size_usd:.0f}/leg | Leverage: {self.config.hedge_leverage}x | Margin: ~${self.config.size_usd*2/self.config.hedge_leverage:.0f}
    - HOLD TIME: {self.config.hedge_hold_hours}h/cycle ({cycles_str} cycles)
    - Exit on fee: {str(self.config.hedge_exit_on_fee):5} | Rest: {self.config.hedge_rest_hours}h
    - Lighter: {'OK' if self.lighter.connected else 'NO'} | HL Auto + Lighter Auto
    - Hold -> Cover Fee -> Close -> Rest -> Re-enter
============================================================
  RISK: High leverage = High liquidation risk!
============================================================

Commands:
  1 or spam  - Start spam mode
  2 or hedge - Start hedge mode  
  close_all  - Close all positions (HL + Lighter)
  stop       - Stop current mode
  status     - Show stats
  quit       - Exit
        """)
        
        while True:
            try:
                cmd = input(f"\n[MODE: {self.mode.upper()}] > ").strip().lower()
                
                if cmd in ("1", "spam"):
                    if self.mode != "off":
                        print("Stop current mode first!")
                        continue
                    try:
                        self.mode = "spam"
                        await self.run_spam_mode()
                    except Exception as e:
                        logger.error(f"Spam mode error: {e}")
                        self.mode = "off"
                    
                elif cmd in ("2", "hedge"):
                    if self.mode != "off":
                        print("Stop current mode first!")
                        continue
                    try:
                        self.mode = "hedge"
                        await self.run_hedge_mode()
                    except Exception as e:
                        logger.error(f"Hedge mode error: {e}")
                        self.mode = "off"
                    
                elif cmd == "stop":
                    self.mode = "off"
                    print("Stopping...")
                
                elif cmd == "close_all":
                    logger.info("Closing ALL positions...")
                    self.hl.close_position()
                    if self.lighter.connected:
                        logger.info("  Waiting 15s for Lighter rate limit...")
                        await asyncio.sleep(15)
                        await self.lighter.close_position()
                    logger.info("  ✅ All positions closed.")
                    
                elif cmd == "status":
                    print(f"\nStats: {json.dumps(self.stats, indent=2, default=str)}")
                    
                elif cmd in ("q", "quit", "exit"):
                    self.mode = "off"
                    print("Goodbye!")
                    break
                    
                else:
                    print("Unknown command")
                    
            except KeyboardInterrupt:
                self.mode = "off"
                break
            except Exception as e:
                logger.error(f"Error: {e}")
        
        # Cleanup
        if self.lighter.connected and self.lighter.trader:
            await self.lighter.trader.close()
            logger.info("Lighter connection closed")


# ═══════════════════════════════════════════════════════════════════
# ⚙️ MODE 2 OI HEDGE SETTINGS - TUỲ CHỈNH SỐ CYCLE VÀ HOLD TIME
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # Presets hay dùng:
    # ─────────────────────────────────────────────────────────────────
    # 1. Short hold (day trading):
    #    hedge_hold_hours=4, hedge_cycles=6    → 6 cycle × 4h = 24h
    #
    # 2. Medium hold (balanced):
    #    hedge_hold_hours=8, hedge_cycles=3    → 3 cycle × 8h = 24h
    #
    # 3. Long hold (lazy):
    #    hedge_hold_hours=12, hedge_cycles=2   → 2 cycle × 12h = 24h
    #
    # 4. Farm OI vô hạn:
    #    hedge_hold_hours=8, hedge_cycles=-1   → Chạy mãi
    # ─────────────────────────────────────────────────────────────────
    
    config = TradeConfig(
        size_usd=150.0,           # $150 per leg (HL x10 + Lighter ~1.2x, an toàn)
        spam_leverage=20,          # Mode 1: 20x cho spam
        hedge_leverage=10,         # Mode 2: 10x cho hedge (an toàn)
        
        # 🔄 MODE 1 SPAM SETTINGS:
        spam_rounds=2,              # 🔁 Số lần spam (10, 20, 50... tuỳ ý)
        spam_interval_sec=10.0,     # ⏱️ Delay giữa các lệnh (10-30s tránh rate limit)
        spam_size_range=(100, 400), # 💰 Random size $100-$400 mỗi lệnh (min, max)
        
        # ⏱️ MODE 2 OI SETTINGS:
        hedge_hold_hours=0.1,      # Hold mỗi cycle (4h-12h)
        hedge_auto_reenter=True,   # Auto vào lệnh mới sau khi đóng
        hedge_cycles=2,            # -1=vô hạn | 3=3 cycles | 6=6 cycles
        hedge_rest_hours=0.05,      # Nghỉ giữa các cycle (giờ)
        hedge_exit_on_fee=True,    # ✅ True=đóng khi đủ gỡ phí | False=chờ đủ profit_target
    )
    
    bot = Bot2(config)
    asyncio.run(bot.run())